import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
import os
from datetime import datetime
import json
from jinja2 import Template
import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

class DataScribeReportGenerator:
    """
    Report Generation Engine for DataScribe MVP
    Creates comprehensive reports in HTML, PDF, and Excel formats
    """
    
    def __init__(self, df: pd.DataFrame, analysis_results: Dict, target_col: Optional[str] = None):
        self.df = df.copy()
        self.analysis_results = analysis_results
        self.target_col = target_col
        self.report_data = self._prepare_report_data()
        
    def _prepare_report_data(self) -> Dict:
        """Prepare all data needed for report generation"""
        return {
            'dataset_info': {
                'name': 'Uploaded Dataset',
                'shape': self.df.shape,
                'memory_usage': f"{self.df.memory_usage(deep=True).sum() / 1024**2:.2f} MB",
                'upload_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'columns': {
                    'total': len(self.df.columns),
                    'numerical': len(self.df.select_dtypes(include=[np.number]).columns),
                    'categorical': len(self.df.select_dtypes(include=['object', 'category']).columns),
                    'datetime': len(self.df.select_dtypes(include=['datetime64']).columns)
                }
            },
            'analysis_results': self.analysis_results,
            'target_column': self.target_col
        }
    
    def generate_html_report(self, plot_files: Dict[str, str]) -> str:
        """Generate comprehensive HTML report"""
        html_template = self._get_html_template()
        
        # Prepare data for template
        template_data = {
            **self.report_data,
            'plot_files': plot_files,
            'generation_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Render template
        template = Template(html_template)
        html_content = template.render(**template_data)
        
        # Save HTML file
        filename = f"datascribe_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return filename
    
    def generate_pdf_report(self, plot_files: Dict[str, str]) -> str:
        """Generate PDF report from HTML"""
        try:
            # First generate HTML
            html_file = self.generate_html_report(plot_files)
            
            # Convert to PDF
            pdf_filename = html_file.replace('.html', '.pdf')
            
            # Read HTML content
            with open(html_file, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # Convert to PDF
            weasyprint.HTML(string=html_content).write_pdf(pdf_filename)
            
            return pdf_filename
            
        except Exception as e:
            print(f"PDF generation failed: {e}")
            return f"PDF generation failed: {str(e)}"
    
    def generate_excel_report(self) -> str:
        """Generate comprehensive Excel report with multiple sheets"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_overview_sheet(wb)
            self._create_data_quality_sheet(wb)
            self._create_statistics_sheet(wb)
            self._create_correlations_sheet(wb)
            self._create_insights_sheet(wb)
            
            # Save workbook
            filename = f"datascribe_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            return filename
            
        except Exception as e:
            print(f"Excel generation failed: {e}")
            return f"Excel generation failed: {str(e)}"
    
    def _create_overview_sheet(self, wb: Workbook):
        """Create overview sheet"""
        ws = wb.create_sheet("Dataset Overview")
        
        # Title
        ws['A1'] = "DataScribe - Dataset Overview"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Dataset information
        ws['A3'] = "Dataset Information"
        ws['A3'].font = Font(size=14, bold=True)
        
        info_data = [
            ["Dataset Name", self.report_data['dataset_info']['name']],
            ["Rows", self.report_data['dataset_info']['shape'][0]],
            ["Columns", self.report_data['dataset_info']['shape'][1]],
            ["Memory Usage", self.report_data['dataset_info']['memory_usage']],
            ["Upload Time", self.report_data['dataset_info']['upload_time']],
            ["Target Column", self.target_col if self.target_col else "Not specified"]
        ]
        
        for i, (label, value) in enumerate(info_data):
            ws[f'A{5+i}'] = label
            ws[f'B{5+i}'] = value
            ws[f'A{5+i}'].font = Font(bold=True)
        
        # Column types
        ws['A12'] = "Column Types Distribution"
        ws['A12'].font = Font(size=14, bold=True)
        
        col_types_data = [
            ["Numerical Columns", self.report_data['dataset_info']['columns']['numerical']],
            ["Categorical Columns", self.report_data['dataset_info']['columns']['categorical']],
            ["Datetime Columns", self.report_data['dataset_info']['columns']['datetime']]
        ]
        
        for i, (label, value) in enumerate(col_types_data):
            ws[f'A{14+i}'] = label
            ws[f'B{14+i}'] = value
            ws[f'A{14+i}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_data_quality_sheet(self, wb: Workbook):
        """Create data quality sheet"""
        ws = wb.create_sheet("Data Quality")
        
        # Title
        ws['A1'] = "Data Quality Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        
        quality_data = self.report_data['analysis_results'].get('data_quality', {})
        
        # Quality score
        ws['A3'] = "Overall Data Quality Score"
        ws['A3'].font = Font(size=14, bold=True)
        ws['B3'] = f"{quality_data.get('data_quality_score', 100):.1f}/100"
        ws['B3'].font = Font(size=14, bold=True, color="00B050" if quality_data.get('data_quality_score', 100) >= 80 else "FF0000")
        
        # Missing values
        ws['A5'] = "Missing Values Analysis"
        ws['A5'].font = Font(size=14, bold=True)
        
        missing_data = quality_data.get('missing_values', {})
        if missing_data.get('total_missing', 0) > 0:
            ws['A7'] = "Column"
            ws['B7'] = "Missing Count"
            ws['C7'] = "Missing %"
            ws['A7'].font = Font(bold=True)
            ws['B7'].font = Font(bold=True)
            ws['C7'].font = Font(bold=True)
            
            missing_counts = missing_data.get('count', {})
            missing_percentages = missing_data.get('percentage', {})
            
            row = 8
            for col, count in missing_counts.items():
                if count > 0:
                    ws[f'A{row}'] = col
                    ws[f'B{row}'] = count
                    ws[f'C{row}'] = f"{missing_percentages.get(col, 0):.2f}%"
                    row += 1
        else:
            ws['A7'] = "No missing values found in the dataset"
            ws['A7'].font = Font(italic=True, color="00B050")
        
        # Duplicates
        ws['A15'] = "Duplicate Analysis"
        ws['A15'].font = Font(size=14, bold=True)
        
        duplicate_info = quality_data.get('duplicates', {})
        ws['A17'] = "Duplicate Rows"
        ws['B17'] = duplicate_info.get('count', 0)
        ws['A18'] = "Duplicate Percentage"
        ws['B18'] = f"{duplicate_info.get('percentage', 0):.2f}%"
        
        # Constant columns
        ws['A20'] = "Constant Columns"
        ws['A20'].font = Font(size=14, bold=True)
        
        constant_cols = quality_data.get('constant_columns', [])
        if constant_cols:
            for i, col in enumerate(constant_cols):
                ws[f'A{22+i}'] = col
        else:
            ws['A22'] = "No constant columns found"
            ws['A22'].font = Font(italic=True, color="00B050")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistics_sheet(self, wb: Workbook):
        """Create statistics sheet"""
        ws = wb.create_sheet("Statistical Summary")
        
        # Title
        ws['A1'] = "Statistical Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        stats_data = self.report_data['analysis_results'].get('statistics', {})
        
        # Numerical statistics
        if 'numerical' in stats_data:
            ws['A3'] = "Numerical Columns Statistics"
            ws['A3'].font = Font(size=14, bold=True)
            
            numerical_stats = stats_data['numerical']
            # Get unique column names (remove _skewness, _kurtosis suffixes)
            base_columns = set()
            for key in numerical_stats.keys():
                if not key.endswith(('_skewness', '_kurtosis')):
                    base_columns.add(key)
            
            if base_columns:
                # Create statistics table
                stats_columns = ['Column', 'Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max', 'Skewness', 'Kurtosis']
                for i, col in enumerate(stats_columns):
                    ws[f'{chr(65+i)}5'] = col
                    ws[f'{chr(65+i)}5'].font = Font(bold=True)
                
                row = 6
                for col in sorted(base_columns):
                    ws[f'A{row}'] = col
                    ws[f'B{row}'] = numerical_stats.get(f'{col}_count', 'N/A')
                    ws[f'C{row}'] = numerical_stats.get(f'{col}_mean', 'N/A')
                    ws[f'D{row}'] = numerical_stats.get(f'{col}_std', 'N/A')
                    ws[f'E{row}'] = numerical_stats.get(f'{col}_min', 'N/A')
                    ws[f'F{row}'] = numerical_stats.get(f'{col}_25%', 'N/A')
                    ws[f'G{row}'] = numerical_stats.get(f'{col}_50%', 'N/A')
                    ws[f'H{row}'] = numerical_stats.get(f'{col}_75%', 'N/A')
                    ws[f'I{row}'] = numerical_stats.get(f'{col}_max', 'N/A')
                    ws[f'J{row}'] = numerical_stats.get(f'{col}_skewness', 'N/A')
                    ws[f'K{row}'] = numerical_stats.get(f'{col}_kurtosis', 'N/A')
                    row += 1
        
        # Categorical statistics
        if 'categorical' in stats_data:
            start_row = 20
            ws[f'A{start_row}'] = "Categorical Columns Statistics"
            ws[f'A{start_row}'].font = Font(size=14, bold=True)
            
            categorical_stats = stats_data['categorical']
            row = start_row + 2
            
            for col, stats in categorical_stats.items():
                ws[f'A{row}'] = f"Column: {col}"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Unique Count"
                ws[f'B{row}'] = stats.get('unique_count', 'N/A')
                row += 1
                
                ws[f'A{row}'] = "Missing Count"
                ws[f'B{row}'] = stats.get('missing_count', 'N/A')
                row += 1
                
                ws[f'A{row}'] = "Top Values"
                row += 1
                
                top_values = stats.get('top_values', {})
                for i, (value, count) in enumerate(list(top_values.items())[:5]):
                    ws[f'B{row}'] = f"{value}: {count}"
                    row += 1
                
                row += 2  # Add space between columns
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_correlations_sheet(self, wb: Workbook):
        """Create correlations sheet"""
        ws = wb.create_sheet("Correlations")
        
        # Title
        ws['A1'] = "Correlation Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        bivariate_data = self.report_data['analysis_results'].get('bivariate', {})
        
        if 'correlations' in bivariate_data:
            corr_data = bivariate_data['correlations']
            
            # High correlations
            ws['A3'] = "High Correlations (|r| > 0.7)"
            ws['A3'].font = Font(size=14, bold=True)
            
            high_corrs = corr_data.get('high_correlations', [])
            if high_corrs:
                ws['A5'] = "Feature 1"
                ws['B5'] = "Feature 2"
                ws['C5'] = "Correlation"
                ws['A5'].font = Font(bold=True)
                ws['B5'].font = Font(bold=True)
                ws['C5'].font = Font(bold=True)
                
                for i, pair in enumerate(high_corrs[:10]):  # Top 10
                    ws[f'A{6+i}'] = pair['feature1']
                    ws[f'B{6+i}'] = pair['feature2']
                    ws[f'C{6+i}'] = pair['correlation']
                    
                    # Color code based on correlation strength
                    if abs(pair['correlation']) > 0.8:
                        ws[f'C{6+i}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws[f'C{6+i}'].font = Font(color="FFFFFF", bold=True)
                    elif abs(pair['correlation']) > 0.7:
                        ws[f'C{6+i}'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                        ws[f'C{6+i}'].font = Font(color="FFFFFF", bold=True)
            else:
                ws['A5'] = "No high correlations found (threshold: 0.7)"
                ws['A5'].font = Font(italic=True, color="00B050")
            
            # Average correlation
            ws['A20'] = "Average Correlation"
            ws['A20'].font = Font(size=14, bold=True)
            ws['B20'] = f"{corr_data.get('avg_correlation', 0):.4f}"
            
            # Correlation matrix (simplified)
            ws['A22'] = "Correlation Matrix (Numerical Features)"
            ws['A22'].font = Font(size=14, bold=True)
            
            corr_matrix = pd.DataFrame(corr_data['correlation_matrix'])
            if not corr_matrix.empty:
                # Add column headers
                for i, col in enumerate(corr_matrix.columns):
                    ws[f'{chr(66+i)}24'] = col
                    ws[f'{chr(66+i)}24'].font = Font(bold=True)
                
                # Add row headers and values
                for i, row_idx in enumerate(corr_matrix.index):
                    ws[f'A{25+i}'] = row_idx
                    ws[f'A{25+i}'].font = Font(bold=True)
                    
                    for j, col in enumerate(corr_matrix.columns):
                        cell_value = corr_matrix.loc[row_idx, col]
                        ws[f'{chr(66+j)}{25+i}'] = f"{cell_value:.3f}"
                        
                        # Color code based on correlation strength
                        if abs(cell_value) > 0.8:
                            ws[f'{chr(66+j)}{25+i}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            ws[f'{chr(66+j)}{25+i}'].font = Font(color="FFFFFF", bold=True)
                        elif abs(cell_value) > 0.6:
                            ws[f'{chr(66+j)}{25+i}'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                            ws[f'{chr(66+j)}{25+i}'].font = Font(color="FFFFFF", bold=True)
        else:
            ws['A3'] = "No correlation data available"
            ws['A3'].font = Font(italic=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_insights_sheet(self, wb: Workbook):
        """Create insights and recommendations sheet"""
        ws = wb.create_sheet("Insights & Recommendations")
        
        # Title
        ws['A1'] = "Insights & Recommendations"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        
        insights_data = self.report_data['analysis_results'].get('insights', {})
        
        # Data quality insights
        ws['A3'] = "Data Quality Insights"
        ws['A3'].font = Font(size=14, bold=True)
        
        quality_insights = insights_data.get('data_quality_insights', [])
        if quality_insights:
            for i, insight in enumerate(quality_insights):
                ws[f'A{5+i}'] = f"• {insight}"
        else:
            ws['A5'] = "No data quality issues identified"
            ws['A5'].font = Font(italic=True, color="00B050")
        
        # Feature insights
        start_row = 10
        ws[f'A{start_row}'] = "Feature Insights"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        
        feature_insights = insights_data.get('feature_insights', [])
        if feature_insights:
            for i, insight in enumerate(feature_insights):
                ws[f'A{start_row+2+i}'] = f"• {insight}"
        else:
            ws[f'A{start_row+2}'] = "No specific feature insights"
            ws[f'A{start_row+2}'].font = Font(italic=True)
        
        # Recommendations
        rec_start_row = 20
        ws[f'A{rec_start_row}'] = "Recommendations"
        ws[f'A{rec_start_row}'].font = Font(size=14, bold=True)
        
        recommendations = insights_data.get('recommendations', [])
        if recommendations:
            for i, rec in enumerate(recommendations):
                ws[f'A{rec_start_row+2+i}'] = f"• {rec}"
        else:
            ws[f'A{rec_start_row+2}'] = "No specific recommendations at this time"
            ws[f'A{rec_start_row+2}'].font = Font(italic=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _get_html_template(self) -> str:
        """Get the HTML template for the report"""
        return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DataScribe Report - {{ dataset_info.name }}</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            background-color: #f8f9fa;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: white;
            box-shadow: 0 0 20px rgba(0,0,0,0.1);
        }
        
        .header {
            text-align: center;
            padding: 40px 0;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 10px;
            margin-bottom: 30px;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 300;
        }
        
        .header .subtitle {
            font-size: 1.2em;
            opacity: 0.9;
            font-weight: 300;
        }
        
        .header .timestamp {
            margin-top: 20px;
            font-size: 0.9em;
            opacity: 0.8;
        }
        
        .section {
            margin-bottom: 40px;
            padding: 25px;
            border-radius: 8px;
            background-color: white;
            border-left: 5px solid #667eea;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
        }
        
        .section h2 {
            color: #2c3e50;
            margin-bottom: 20px;
            font-size: 1.8em;
            border-bottom: 2px solid #ecf0f1;
            padding-bottom: 10px;
        }
        
        .section h3 {
            color: #34495e;
            margin: 20px 0 15px 0;
            font-size: 1.4em;
        }
        
        .info-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        
        .info-card {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        
        .info-card .number {
            font-size: 2em;
            font-weight: bold;
            margin-bottom: 10px;
        }
        
        .info-card .label {
            font-size: 0.9em;
            opacity: 0.9;
        }
        
        .stats-table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background-color: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        
        .stats-table th {
            background-color: #667eea;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }
        
        .stats-table td {
            padding: 12px 15px;
            border-bottom: 1px solid #ecf0f1;
        }
        
        .stats-table tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        
        .stats-table tr:hover {
            background-color: #e3f2fd;
        }
        
        .plot-container {
            text-align: center;
            margin: 30px 0;
            padding: 20px;
            background-color: #f8f9fa;
            border-radius: 8px;
        }
        
        .plot-container img {
            max-width: 100%;
            height: auto;
            border-radius: 8px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        
        .plot-container h4 {
            margin-bottom: 15px;
            color: #2c3e50;
            font-size: 1.2em;
        }
        
        .insights {
            background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);
            padding: 25px;
            border-radius: 8px;
            margin: 20px 0;
        }
        
        .insights h3 {
            color: #2c3e50;
            margin-bottom: 15px;
        }
        
        .insights ul {
            list-style: none;
            padding-left: 0;
        }
        
        .insights li {
            padding: 8px 0;
            border-bottom: 1px solid rgba(255,255,255,0.3);
        }
        
        .insights li:before {
            content: "💡";
            margin-right: 10px;
        }
        
        .footer {
            text-align: center;
            padding: 30px 0;
            color: #7f8c8d;
            border-top: 1px solid #ecf0f1;
            margin-top: 40px;
        }
        
        .quality-score {
            display: inline-block;
            padding: 10px 20px;
            background: linear-gradient(135deg, #56ab2f 0%, #a8e6cf 100%);
            color: white;
            border-radius: 25px;
            font-size: 1.2em;
            font-weight: bold;
            margin: 20px 0;
        }
        
        .warning {
            background-color: #fff3cd;
            border: 1px solid #ffeaa7;
            color: #856404;
            padding: 15px;
            border-radius: 5px;
            margin: 15px 0;
        }
        
        .success {
            background-color: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
            padding: 15px;
            border-radius: 5px;
            margin: 15px 0;
        }
        
        @media (max-width: 768px) {
            .container {
                padding: 10px;
            }
            
            .header h1 {
                font-size: 2em;
            }
            
            .info-grid {
                grid-template-columns: 1fr;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>DataScribe</h1>
            <div class="subtitle">Democratizing Data Analysis: Automated EDA with Human-Readable Insights</div>
            <div class="timestamp">Report Generated: {{ generation_time }}</div>
        </div>
        
        <div class="section">
            <h2>Dataset Overview</h2>
            <div class="info-grid">
                <div class="info-card">
                    <div class="number">{{ dataset_info.shape[0] | number }}</div>
                    <div class="label">Rows</div>
                </div>
                <div class="info-card">
                    <div class="number">{{ dataset_info.shape[1] | number }}</div>
                    <div class="label">Columns</div>
                </div>
                <div class="info-card">
                    <div class="number">{{ dataset_info.memory_usage }}</div>
                    <div class="label">Memory Usage</div>
                </div>
                <div class="info-card">
                    <div class="number">{{ dataset_info.columns.total }}</div>
                    <div class="label">Total Columns</div>
                </div>
            </div>
            
            <h3>Column Types</h3>
            <table class="stats-table">
                <thead>
                    <tr>
                        <th>Type</th>
                        <th>Count</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>Numerical</td>
                        <td>{{ dataset_info.columns.numerical }}</td>
                        <td>{{ "%.1f" | format(dataset_info.columns.numerical / dataset_info.columns.total * 100) }}%</td>
                    </tr>
                    <tr>
                        <td>Categorical</td>
                        <td>{{ dataset_info.columns.categorical }}</td>
                        <td>{{ "%.1f" | format(dataset_info.columns.categorical / dataset_info.columns.total * 100) }}%</td>
                    </tr>
                    <tr>
                        <td>Datetime</td>
                        <td>{{ dataset_info.columns.datetime }}</td>
                        <td>{{ "%.1f" | format(dataset_info.columns.datetime / dataset_info.columns.total * 100) }}%</td>
                    </tr>
                </tbody>
            </table>
            
            {% if target_column %}
            <h3>Target Variable</h3>
            <div class="success">
                <strong>Target Column:</strong> {{ target_column }}
            </div>
            {% endif %}
        </div>
        
        <div class="section">
            <h2>Data Quality Analysis</h2>
            {% set quality = analysis_results.data_quality %}
            {% if quality %}
                <div class="quality-score">
                    Data Quality Score: {{ "%.1f" | format(quality.data_quality_score) }}/100
                </div>
                
                {% if quality.missing_values.total_missing > 0 %}
                <div class="warning">
                    <strong>Missing Values Detected:</strong> {{ quality.missing_values.total_missing }} total missing values across the dataset.
                </div>
                {% else %}
                <div class="success">
                    <strong>No Missing Values:</strong> Your dataset is complete!
                </div>
                {% endif %}
                
                {% if quality.duplicates.count > 0 %}
                <div class="warning">
                    <strong>Duplicates Found:</strong> {{ quality.duplicates.count }} duplicate rows ({{ "%.2f" | format(quality.duplicates.percentage) }}% of dataset)
                </div>
                {% else %}
                <div class="success">
                    <strong>No Duplicates:</strong> Your dataset contains unique records.
                </div>
                {% endif %}
                
                {% if quality.constant_columns %}
                <div class="warning">
                    <strong>Constant Columns:</strong> {{ quality.constant_columns | length }} columns have no variation in values.
                </div>
                {% endif %}
            {% endif %}
        </div>
        
        {% if plot_files.overview %}
        <div class="section">
            <h2>Dataset Overview Visualizations</h2>
            <div class="plot-container">
                <h4>Dataset Overview</h4>
                <img src="{{ plot_files.overview }}" alt="Dataset Overview">
            </div>
        </div>
        {% endif %}
        
        {% if plot_files.data_quality %}
        <div class="section">
            <h2>Data Quality Visualizations</h2>
            <div class="plot-container">
                <h4>Data Quality Analysis</h4>
                <img src="{{ plot_files.data_quality }}" alt="Data Quality Analysis">
            </div>
        </div>
        {% endif %}
        
        {% if plot_files.univariate %}
        <div class="section">
            <h2>Univariate Analysis</h2>
            <div class="plot-container">
                <h4>Feature Distributions</h4>
                <img src="{{ plot_files.univariate }}" alt="Univariate Analysis">
            </div>
        </div>
        {% endif %}
        
        {% if plot_files.bivariate %}
        <div class="section">
            <h2>Bivariate Analysis</h2>
            <div class="plot-container">
                <h4>Correlation Analysis</h4>
                <img src="{{ plot_files.bivariate }}" alt="Bivariate Analysis">
            </div>
        </div>
        {% endif %}
        
        {% if plot_files.multivariate %}
        <div class="section">
            <h2>Multivariate Analysis</h2>
            <div class="plot-container">
                <h4>Advanced Analysis</h4>
                <img src="{{ plot_files.multivariate }}" alt="Multivariate Analysis">
            </div>
        </div>
        {% endif %}
        
        {% if target_column and plot_files.target_analysis %}
        <div class="section">
            <h2>Target Variable Analysis</h2>
            <div class="plot-container">
                <h4>Target Analysis</h4>
                <img src="{{ plot_files.target_analysis }}" alt="Target Analysis">
            </div>
        </div>
        {% endif %}
        
        {% if analysis_results.insights %}
        <div class="section">
            <h2>Insights & Recommendations</h2>
            {% set insights = analysis_results.insights %}
            
            {% if insights.data_quality_insights %}
            <h3>Data Quality Insights</h3>
            <div class="insights">
                <ul>
                    {% for insight in insights.data_quality_insights %}
                    <li>{{ insight }}</li>
                    {% endfor %}
                </ul>
            </div>
            {% endif %}
            
            {% if insights.feature_insights %}
            <h3>Feature Insights</h3>
            <div class="insights">
                <ul>
                    {% for insight in insights.feature_insights %}
                    <li>{{ insight }}</li>
                    {% endfor %}
                </ul>
            </div>
            {% endif %}
            
            {% if insights.recommendations %}
            <h3>Recommendations</h3>
            <div class="insights">
                <ul>
                    {% for rec in insights.recommendations %}
                    <li>{{ rec }}</li>
                    {% endfor %}
                </ul>
            </div>
            {% endif %}
        </div>
        {% endif %}
        
        <div class="footer">
            <p>Generated by DataScribe - AI-Powered EDA Platform</p>
            <p>Report ID: {{ generation_time | replace(" ", "_") | replace(":", "-") }}</p>
        </div>
    </div>
</body>
</html>
        """

def generate_reports(df: pd.DataFrame, analysis_results: Dict, plot_files: Dict[str, str], target_col: Optional[str] = None) -> Dict[str, str]:
    """
    Main function to generate all report formats
    """
    generator = DataScribeReportGenerator(df, analysis_results, target_col)
    
    reports = {}
    
    # Generate HTML report
    reports['html'] = generator.generate_html_report(plot_files)
    
    # Generate PDF report
    reports['pdf'] = generator.generate_pdf_report(plot_files)
    
    # Generate Excel report
    reports['excel'] = generator.generate_excel_report()
    
    return reports
